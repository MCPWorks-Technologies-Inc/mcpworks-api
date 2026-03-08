#!/usr/bin/env python3
"""Sandbox smoke test — validate all sandbox packages work inside nsjail.

Imports every package from PACKAGE_REGISTRY and performs a real operation
(not just import) to exercise native code, shared libraries, and syscalls.

Usage:
    python3 smoketest.py           # Human-readable output
    python3 smoketest.py --json    # JSON output for CI parsing

Exit codes:
    0 = all packages passed
    1 = one or more packages failed
"""

import io
import json
import sys
import time
import traceback

# ---------------------------------------------------------------------------
# Package test definitions
#
# Each entry: (import_name, description, test_callable)
# test_callable receives no args, returns a truthy value on success.
# If it raises, the package is marked as failed.
# ---------------------------------------------------------------------------

TESTS: list[tuple[str, str, object]] = []


def _reg(import_name: str, desc: str, fn):
    TESTS.append((import_name, desc, fn))


# ── HTTP & Networking ──


def _test_requests():
    import requests

    session = requests.Session()
    p = session.prepare_request(requests.Request("GET", "http://example.com"))
    return p.method == "GET"


def _test_httpx():
    import httpx

    r = httpx.Request("GET", "http://example.com")
    return r.method == "GET"


def _test_urllib3():
    import urllib3

    p = urllib3.HTTPConnectionPool("localhost", maxsize=1)
    return p is not None


def _test_aiohttp():
    import aiohttp

    return hasattr(aiohttp, "ClientSession")


def _test_websockets():
    import websockets

    return hasattr(websockets, "connect")


_reg("requests", "Build a prepared request", _test_requests)
_reg("httpx", "Build an httpx request", _test_httpx)
_reg("urllib3", "Create a connection pool", _test_urllib3)
_reg("aiohttp", "Import aiohttp client", _test_aiohttp)
_reg("websockets", "Import websockets", _test_websockets)

# ── Data Formats ──


def _test_pyyaml():
    import yaml

    data = yaml.safe_load("key: value\nlist:\n  - 1\n  - 2")
    return data["key"] == "value" and data["list"] == [1, 2]


def _test_orjson():
    import orjson

    data = {"key": "value", "nums": [1, 2, 3]}
    encoded = orjson.dumps(data)
    decoded = orjson.loads(encoded)
    return decoded == data


def _test_tomli():
    import tomli

    data = tomli.loads('[section]\nkey = "value"')
    return data["section"]["key"] == "value"


def _test_tomli_w():
    import tomli_w

    result = tomli_w.dumps({"key": "value"})
    return 'key = "value"' in result


def _test_xmltodict():
    import xmltodict

    data = xmltodict.parse("<root><item>hello</item></root>")
    return data["root"]["item"] == "hello"


def _test_msgpack():
    import msgpack

    data = {"key": "value", "nums": [1, 2, 3]}
    packed = msgpack.packb(data)
    unpacked = msgpack.unpackb(packed)
    return unpacked == data


_reg("yaml", "Parse YAML", _test_pyyaml)
_reg("orjson", "Round-trip JSON", _test_orjson)
_reg("tomli", "Parse TOML", _test_tomli)
_reg("tomli_w", "Write TOML", _test_tomli_w)
_reg("xmltodict", "Parse XML to dict", _test_xmltodict)
_reg("msgpack", "Round-trip msgpack", _test_msgpack)

# ── Data Validation ──


def _test_pydantic():
    from pydantic import BaseModel

    class M(BaseModel):
        name: str
        age: int

    m = M(name="test", age=42)
    return m.name == "test" and m.age == 42


def _test_attrs():
    import attr

    @attr.s(auto_attribs=True)
    class C:
        x: int = 0

    c = C(x=42)
    return c.x == 42


def _test_jsonschema():
    import jsonschema

    schema = {"type": "object", "properties": {"name": {"type": "string"}}}
    jsonschema.validate({"name": "test"}, schema)
    return True


_reg("pydantic", "Validate a model", _test_pydantic)
_reg("attr", "Create attrs class", _test_attrs)
_reg("jsonschema", "Validate JSON schema", _test_jsonschema)

# ── Text & Content Processing ──


def _test_bs4():
    from bs4 import BeautifulSoup

    soup = BeautifulSoup("<html><body><p>Hello</p></body></html>", "html.parser")
    return soup.find("p").text == "Hello"


def _test_lxml():
    from lxml import etree

    root = etree.fromstring(b"<root><child>text</child></root>")
    result = etree.tostring(root, encoding="unicode")
    return "<child>text</child>" in result


def _test_markdownify():
    import markdownify

    result = markdownify.markdownify("<h1>Title</h1><p>Text</p>")
    return "Title" in result


def _test_markdown():
    import markdown

    result = markdown.markdown("# Title\n\nParagraph")
    return "<h1>Title</h1>" in result


def _test_html2text():
    import html2text

    h = html2text.HTML2Text()
    result = h.handle("<h1>Title</h1><p>Text</p>")
    return "Title" in result


def _test_chardet():
    import chardet

    result = chardet.detect(b"Hello world")
    return result["encoding"] is not None


def _test_slugify():
    from slugify import slugify

    return slugify("Hello World!") == "hello-world"


def _test_jinja2():
    from jinja2 import Template

    t = Template("Hello {{ name }}!")
    return t.render(name="World") == "Hello World!"


def _test_regex():
    import regex

    m = regex.search(r"\p{L}+", "Hello123")
    return m.group() == "Hello"


_reg("bs4", "Parse HTML", _test_bs4)
_reg("lxml", "Parse + serialize XML", _test_lxml)
_reg("markdownify", "HTML to markdown", _test_markdownify)
_reg("markdown", "Markdown to HTML", _test_markdown)
_reg("html2text", "HTML to text", _test_html2text)
_reg("chardet", "Detect encoding", _test_chardet)
_reg("slugify", "Slugify string", _test_slugify)
_reg("jinja2", "Render template", _test_jinja2)
_reg("regex", "Unicode regex", _test_regex)

# ── Date & Time ──


def _test_dateutil():
    from dateutil import parser as dp

    dt = dp.parse("2025-01-15T10:30:00Z")
    return dt.year == 2025 and dt.month == 1


def _test_pytz():
    import pytz

    tz = pytz.timezone("US/Eastern")
    return tz is not None


def _test_arrow():
    import arrow

    a = arrow.get("2025-01-15", "YYYY-MM-DD")
    return a.year == 2025


_reg("dateutil", "Parse date string", _test_dateutil)
_reg("pytz", "Load timezone", _test_pytz)
_reg("arrow", "Parse with arrow", _test_arrow)

# ── Data Science ──


def _test_numpy():
    import numpy as np

    a = np.random.rand(50, 50)
    b = np.random.rand(50, 50)
    c = a @ b  # Matrix multiply — exercises BLAS/OpenBLAS (mbind)
    return c.shape == (50, 50)


def _test_pandas():
    import pandas as pd

    df = pd.DataFrame({"a": [1, 2, 3], "b": [4, 5, 6]})
    result = df.groupby("a").sum()
    return len(result) == 3


def _test_scipy():
    from scipy import optimize

    result = optimize.minimize_scalar(lambda x: (x - 3) ** 2)
    return abs(result.x - 3.0) < 0.01


def _test_sklearn():
    import numpy as np
    from sklearn.linear_model import LinearRegression

    X = np.array([[1], [2], [3], [4]])
    y = np.array([2, 4, 6, 8])
    model = LinearRegression().fit(X, y)
    pred = model.predict([[5]])
    return abs(pred[0] - 10.0) < 0.1


def _test_sympy():
    from sympy import expand, symbols

    x = symbols("x")
    expr = expand((x + 1) ** 3)
    return "x**3" in str(expr)


def _test_statsmodels():
    import numpy as np
    import statsmodels.api as sm

    X = np.array([[1], [2], [3], [4], [5]], dtype=float)
    y = np.array([2.1, 3.9, 6.2, 7.8, 10.1])
    X = sm.add_constant(X)
    model = sm.OLS(y, X).fit()
    return model.rsquared > 0.9


_reg("numpy", "50x50 matrix multiply (BLAS)", _test_numpy)
_reg("pandas", "DataFrame groupby", _test_pandas)
_reg("scipy", "Scalar optimization", _test_scipy)
_reg("sklearn", "Linear regression fit+predict", _test_sklearn)
_reg("sympy", "Symbolic expansion", _test_sympy)
_reg("statsmodels", "OLS regression", _test_statsmodels)

# ── Visualization ──


def _test_matplotlib():
    import matplotlib

    matplotlib.use("Agg")
    import matplotlib.pyplot as plt

    fig, ax = plt.subplots()
    ax.plot([1, 2, 3], [1, 4, 9])
    buf = io.BytesIO()
    fig.savefig(buf, format="png")
    plt.close(fig)
    return buf.tell() > 100  # PNG should be at least a few hundred bytes


def _test_pillow():
    from PIL import Image

    img = Image.new("RGB", (100, 100), color="red")
    buf = io.BytesIO()
    img.save(buf, format="PNG")
    return buf.tell() > 50


_reg("matplotlib", "Render plot to PNG", _test_matplotlib)
_reg("PIL", "Create + save image", _test_pillow)

# ── AI & LLM ──


def _test_openai():
    import openai

    client = openai.OpenAI(api_key="sk-test", base_url="http://localhost:1")
    return client is not None


def _test_anthropic():
    import anthropic

    client = anthropic.Anthropic(api_key="sk-ant-test")
    return client is not None


def _test_tiktoken():
    import tiktoken

    enc = tiktoken.get_encoding("cl100k_base")
    tokens = enc.encode("Hello world")
    return len(tokens) > 0


def _test_cohere():
    import cohere

    return hasattr(cohere, "ClientV2") or hasattr(cohere, "Client")


_reg("openai", "Create OpenAI client", _test_openai)
_reg("anthropic", "Create Anthropic client", _test_anthropic)
_reg("tiktoken", "Encode tokens (cl100k_base)", _test_tiktoken)
_reg("cohere", "Import cohere client", _test_cohere)

# ── Cloud & SaaS APIs ──


def _test_boto3():
    import botocore.session

    session = botocore.session.get_session()
    return session is not None


def _test_stripe():
    import stripe

    stripe.api_key = "sk_test_fake"
    return hasattr(stripe, "Customer")


def _test_sendgrid():
    from sendgrid import SendGridAPIClient

    return SendGridAPIClient is not None


def _test_twilio():
    from twilio.rest import Client as TwilioClient

    return TwilioClient is not None


def _test_gcs():
    from google.cloud import storage

    return hasattr(storage, "Client")


_reg("boto3", "Create botocore session", _test_boto3)
_reg("stripe", "Import stripe SDK", _test_stripe)
_reg("sendgrid", "Import SendGrid client", _test_sendgrid)
_reg("twilio", "Import Twilio client", _test_twilio)
_reg("google.cloud.storage", "Import GCS client", _test_gcs)

# ── File Formats ──


def _test_tabulate():
    from tabulate import tabulate

    result = tabulate([["Alice", 24], ["Bob", 30]], headers=["Name", "Age"])
    return "Alice" in result and "Bob" in result


def _test_feedparser():
    import feedparser

    feed = feedparser.parse("<rss><channel><title>Test</title></channel></rss>")
    return feed.feed.get("title") == "Test"


def _test_openpyxl():
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws["A1"] = "Hello"
    buf = io.BytesIO()
    wb.save(buf)
    return buf.tell() > 100


def _test_xlsxwriter():
    import xlsxwriter

    buf = io.BytesIO()
    wb = xlsxwriter.Workbook(buf, {"in_memory": True})
    ws = wb.add_worksheet()
    ws.write(0, 0, "Hello")
    wb.close()
    return buf.tell() > 100


def _test_docx():
    from docx import Document

    doc = Document()
    doc.add_paragraph("Hello World")
    buf = io.BytesIO()
    doc.save(buf)
    return buf.tell() > 100


def _test_pypdf():
    import pypdf

    return hasattr(pypdf, "PdfReader")


_reg("tabulate", "Format table", _test_tabulate)
_reg("feedparser", "Parse RSS feed", _test_feedparser)
_reg("openpyxl", "Create Excel workbook", _test_openpyxl)
_reg("xlsxwriter", "Write Excel file", _test_xlsxwriter)
_reg("docx", "Create Word document", _test_docx)
_reg("pypdf", "Import pypdf", _test_pypdf)

# ── Crypto & Security ──


def _test_cryptography():
    from cryptography.hazmat.primitives import hashes
    from cryptography.hazmat.primitives.asymmetric import padding, rsa

    key = rsa.generate_private_key(public_exponent=65537, key_size=2048)
    message = b"test message"
    sig = key.sign(message, padding.PKCS1v15(), hashes.SHA256())
    key.public_key().verify(sig, message, padding.PKCS1v15(), hashes.SHA256())
    return True


def _test_pyjwt():
    import jwt

    token = jwt.encode({"sub": "1234"}, "secret", algorithm="HS256")
    data = jwt.decode(token, "secret", algorithms=["HS256"])
    return data["sub"] == "1234"


def _test_bcrypt():
    import bcrypt

    pw = b"test_password"
    hashed = bcrypt.hashpw(pw, bcrypt.gensalt(rounds=4))
    return bcrypt.checkpw(pw, hashed)


_reg("cryptography", "RSA keygen + sign/verify", _test_cryptography)
_reg("jwt", "JWT encode/decode", _test_pyjwt)
_reg("bcrypt", "Hash + verify password", _test_bcrypt)

# ── Database Clients ──


def _test_psycopg2():
    import psycopg2

    return hasattr(psycopg2, "connect")


def _test_pymongo():
    import pymongo

    return hasattr(pymongo, "MongoClient")


def _test_redis():
    import redis

    return hasattr(redis, "Redis")


_reg("psycopg2", "Import psycopg2", _test_psycopg2)
_reg("pymongo", "Import pymongo", _test_pymongo)
_reg("redis", "Import redis", _test_redis)

# ── Utilities ──


def _test_humanize():
    import humanize

    return humanize.naturalsize(1048576) == "1.0 MB"


def _test_tqdm():
    from tqdm import tqdm

    result = list(tqdm(range(10), disable=True))
    return result == list(range(10))


def _test_rich():
    from rich.console import Console

    console = Console(file=io.StringIO())
    console.print("[bold]Hello[/bold]")
    return True


def _test_typing_extensions():
    import typing_extensions

    return hasattr(typing_extensions, "TypedDict")


_reg("humanize", "Format file size", _test_humanize)
_reg("tqdm", "Iterate with progress", _test_tqdm)
_reg("rich", "Rich console output", _test_rich)
_reg("typing_extensions", "Import typing_extensions", _test_typing_extensions)


# ---------------------------------------------------------------------------
# Runner
# ---------------------------------------------------------------------------


def run_all(json_output: bool = False) -> int:
    results = []
    passed = 0
    failed = 0

    for import_name, desc, fn in TESTS:
        t0 = time.monotonic()
        try:
            ok = fn()
            elapsed = time.monotonic() - t0
            if ok:
                passed += 1
                results.append(
                    {
                        "package": import_name,
                        "description": desc,
                        "status": "pass",
                        "time_ms": round(elapsed * 1000, 1),
                    }
                )
                if not json_output:
                    print(f"  PASS  {import_name:<30s} {desc:<40s} ({elapsed * 1000:.0f}ms)")
            else:
                failed += 1
                results.append(
                    {
                        "package": import_name,
                        "description": desc,
                        "status": "fail",
                        "error": "returned falsy",
                        "time_ms": round(elapsed * 1000, 1),
                    }
                )
                if not json_output:
                    print(f"  FAIL  {import_name:<30s} {desc:<40s} (returned falsy)")
        except Exception as e:
            elapsed = time.monotonic() - t0
            failed += 1
            tb = traceback.format_exc()
            results.append(
                {
                    "package": import_name,
                    "description": desc,
                    "status": "fail",
                    "error": str(e),
                    "traceback": tb,
                    "time_ms": round(elapsed * 1000, 1),
                }
            )
            if not json_output:
                print(f"  FAIL  {import_name:<30s} {desc:<40s} ({e})")

    total = passed + failed
    sys.stderr.write(f"SMOKETEST: {passed}/{total} passed, {failed} failed\n")
    sys.stderr.flush()

    if json_output:
        output = {
            "total": total,
            "passed": passed,
            "failed": failed,
            "results": results,
        }
        json_str = json.dumps(output, indent=2)
        try:
            with open("/sandbox/smoketest-output.json", "w") as f:
                f.write(json_str)
            sys.stderr.write(f"SMOKETEST: wrote {len(json_str)} bytes to output file\n")
        except OSError as e:
            sys.stderr.write(f"SMOKETEST: file write failed: {e}\n")
        sys.stderr.flush()
        try:
            sys.stdout.write(json_str + "\n")
            sys.stdout.flush()
            sys.stderr.write("SMOKETEST: stdout write ok\n")
        except OSError as e:
            sys.stderr.write(f"SMOKETEST: stdout write failed: {e}\n")
        sys.stderr.flush()
    else:
        print(f"\n{'=' * 70}")
        print(f"  {passed}/{total} passed, {failed} failed")
        if failed:
            print("\n  Failed packages:")
            for r in results:
                if r["status"] == "fail":
                    print(f"    - {r['package']}: {r.get('error', 'unknown')}")
        print(f"{'=' * 70}")

    return 1 if failed else 0


if __name__ == "__main__":
    json_output = "--json" in sys.argv
    sys.exit(run_all(json_output=json_output))
